import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
TABLE_TOOL = ROOT / ".opencode" / "tools" / "table.py"


def run_table_tool(request):
    return subprocess.run(
        [sys.executable, str(TABLE_TOOL), json.dumps(request)],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=False,
    )


def make_workbook(path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Audit Output"
    sheet["A1"] = "Template"
    workbook.save(path)


class TableWriteValuesTest(unittest.TestCase):
    def test_writes_value_directly_and_records_metadata(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "audit.xlsx"
            make_workbook(workbook_path)

            result = run_table_tool({
                "action": "write_values",
                "path": str(workbook_path),
                "records": [
                    {
                        "output_sheet": "Audit Output",
                        "output_cell": "D12",
                        "value": "Yes",
                        "sql": "SELECT note_text FROM clinical_notes WHERE patient_code = 'P1'",
                        "explanation": "Note documents a sentinel event, so the field is Yes.",
                    },
                ],
            })
            self.assertEqual(result.returncode, 0, result.stderr)
            payload = json.loads(result.stdout)
            self.assertEqual(payload["cellsWritten"], 1)

            # Value lands directly in the cell (readable with data_only).
            wb = load_workbook(workbook_path, data_only=True)
            self.assertEqual(wb["Audit Output"]["D12"].value, "Yes")
            self.assertNotIn("Evidence", wb.sheetnames)

            # Metadata sidecar exists next to the workbook with the right entry.
            meta_path = Path(directory) / "metadata.json"
            self.assertTrue(meta_path.exists())
            meta = json.loads(meta_path.read_text())
            entry = meta["Audit Output!D12"]
            self.assertEqual(entry["value"], "Yes")
            self.assertIn("clinical_notes", entry["sql"])
            self.assertEqual(entry["explanation"], "Note documents a sentinel event, so the field is Yes.")

    def test_metadata_upserts_across_calls(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "audit.xlsx"
            make_workbook(workbook_path)

            run_table_tool({
                "action": "write_values",
                "path": str(workbook_path),
                "records": [{
                    "output_sheet": "Audit Output", "output_cell": "A2",
                    "value": "P1", "sql": "SELECT code FROM patients WHERE id = 1",
                }],
            })
            run_table_tool({
                "action": "write_values",
                "path": str(workbook_path),
                "records": [{
                    "output_sheet": "Audit Output", "output_cell": "B2",
                    "value": 7.1, "sql": "SELECT hba1c FROM obs WHERE id = 1",
                }],
            })

            meta = json.loads((Path(directory) / "metadata.json").read_text())
            self.assertIn("Audit Output!A2", meta)
            self.assertIn("Audit Output!B2", meta)
            self.assertIsNone(meta["Audit Output!A2"]["explanation"])

    def test_missing_sql_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "audit.xlsx"
            make_workbook(workbook_path)

            result = run_table_tool({
                "action": "write_values",
                "path": str(workbook_path),
                "records": [{"output_sheet": "Audit Output", "output_cell": "A2", "value": "x"}],
            })
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("sql", result.stderr.lower())


if __name__ == "__main__":
    unittest.main()
