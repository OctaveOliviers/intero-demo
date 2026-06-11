import json
import sqlite3
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
POPULATE_TOOL = ROOT / ".opencode" / "tools" / "populate.py"


def run_populate(request):
    return subprocess.run(
        [sys.executable, str(POPULATE_TOOL), json.dumps({"action": "region", **request})],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=False,
    )


def make_db(path: Path):
    conn = sqlite3.connect(path)
    conn.executescript(
        """
        CREATE TABLE patients (mrn TEXT PRIMARY KEY, dob TEXT, last_hba1c REAL);
        INSERT INTO patients VALUES ('M001', '1965-04-12', 7.2);
        INSERT INTO patients VALUES ('M002', '1972-11-30', 6.8);
        INSERT INTO patients VALUES ('M003', '1980-01-05', 8.4);
        """
    )
    conn.commit()
    conn.close()


def make_workbook(path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Patients"
    sheet["A3"] = "MRN"
    sheet["B3"] = "DOB"
    sheet["C3"] = "HbA1c"
    summary = workbook.create_sheet("Summary")
    summary["A3"] = "Total patients"
    workbook.save(path)


class PopulateRegionTest(unittest.TestCase):
    def test_row_per_entity_writes_values_and_metadata(self):
        with tempfile.TemporaryDirectory() as directory:
            db_path = Path(directory) / "emr.sqlite"
            workbook_path = Path(directory) / "audit.xlsx"
            make_db(db_path)
            make_workbook(workbook_path)

            proc = run_populate({
                "workbookPath": str(workbook_path),
                "databasePath": str(db_path),
                "runId": "test-run",
                "region": {
                    "regionId": "R1",
                    "sheet": "Patients",
                    "kind": "row_per_entity",
                    "anchor": {"dataRange": "A4:C10", "rowIdColumn": "mrn"},
                    "fields": [
                        {"col": "A", "resultColumn": "mrn", "sourceTable": "patients", "sourceColumn": "mrn"},
                        {"col": "B", "resultColumn": "dob", "sourceTable": "patients", "sourceColumn": "dob"},
                        {"col": "C", "resultColumn": "last_hba1c", "sourceTable": "patients", "sourceColumn": "last_hba1c"},
                    ],
                },
                "query": {"sql": "SELECT mrn, dob, last_hba1c FROM patients ORDER BY mrn", "parameters": []},
            })
            self.assertEqual(proc.returncode, 0, proc.stderr)
            payload = json.loads(proc.stdout)
            self.assertTrue(payload["ok"], payload)
            self.assertEqual(payload["rowsWritten"], 3)
            self.assertEqual(payload["cellsWritten"], 9)

            # Values land directly in cells (readable with data_only) — no Evidence sheet.
            wb = load_workbook(workbook_path, data_only=True)
            patients = wb["Patients"]
            self.assertEqual(patients["A4"].value, "M001")
            self.assertEqual(patients["C4"].value, 7.2)
            self.assertNotIn("Evidence", wb.sheetnames)

            # Metadata sidecar carries a scoped per-cell SQL.
            meta = json.loads((Path(directory) / "metadata.json").read_text())
            entry = meta["Patients!C4"]
            self.assertEqual(entry["value"], 7.2)
            self.assertIsNone(entry["explanation"])
            self.assertIn("last_hba1c", entry["sql"])
            self.assertIn("M001", entry["sql"])

    def test_column_mismatch_reports_error(self):
        with tempfile.TemporaryDirectory() as directory:
            db_path = Path(directory) / "emr.sqlite"
            workbook_path = Path(directory) / "audit.xlsx"
            make_db(db_path)
            make_workbook(workbook_path)

            proc = run_populate({
                "workbookPath": str(workbook_path),
                "databasePath": str(db_path),
                "runId": "test-run",
                "region": {
                    "regionId": "R1",
                    "sheet": "Patients",
                    "kind": "row_per_entity",
                    "anchor": {"dataRange": "A4:C10"},
                    "fields": [
                        {"col": "A", "resultColumn": "mrn", "sourceTable": "patients", "sourceColumn": "mrn"},
                        {"col": "B", "resultColumn": "age", "sourceTable": "patients", "sourceColumn": "age"},
                    ],
                },
                "query": {"sql": "SELECT mrn, dob FROM patients", "parameters": []},
            })
            payload = json.loads(proc.stdout)
            self.assertFalse(payload["ok"])
            self.assertEqual(payload["errors"][0]["type"], "column_mismatch")
            self.assertIn("age", payload["errors"][0]["missing"])

    def test_non_select_query_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            db_path = Path(directory) / "emr.sqlite"
            workbook_path = Path(directory) / "audit.xlsx"
            make_db(db_path)
            make_workbook(workbook_path)

            proc = run_populate({
                "workbookPath": str(workbook_path),
                "databasePath": str(db_path),
                "runId": "test-run",
                "region": {
                    "regionId": "R1",
                    "sheet": "Patients",
                    "kind": "row_per_entity",
                    "anchor": {"dataRange": "A4:A10"},
                    "fields": [
                        {"col": "A", "resultColumn": "mrn", "sourceTable": "patients", "sourceColumn": "mrn"},
                    ],
                },
                "query": {"sql": "DELETE FROM patients", "parameters": []},
            })
            payload = json.loads(proc.stdout)
            self.assertFalse(payload["ok"])
            self.assertEqual(payload["errors"][0]["type"], "sql_error")

            # workbook unchanged: no Evidence sheet, original cell intact
            workbook = load_workbook(workbook_path)
            self.assertNotIn("Evidence", workbook.sheetnames)

            # database unchanged
            conn = sqlite3.connect(db_path)
            count = conn.execute("SELECT COUNT(*) FROM patients").fetchone()[0]
            conn.close()
            self.assertEqual(count, 3)

    def test_mixed_mode_writes_direct_and_reports_interpret_with_row_mapping(self):
        with tempfile.TemporaryDirectory() as directory:
            db_path = Path(directory) / "emr.sqlite"
            workbook_path = Path(directory) / "audit.xlsx"
            make_db(db_path)
            make_workbook(workbook_path)

            proc = run_populate({
                "workbookPath": str(workbook_path),
                "databasePath": str(db_path),
                "runId": "test-run",
                "region": {
                    "regionId": "R1",
                    "sheet": "Patients",
                    "kind": "row_per_entity",
                    "anchor": {"dataRange": "A4:D10", "rowIdColumn": "mrn"},
                    "fields": [
                        {"col": "A", "mode": "direct", "resultColumn": "mrn", "sourceTable": "patients", "sourceColumn": "mrn"},
                        {"col": "B", "mode": "direct", "resultColumn": "dob", "sourceTable": "patients", "sourceColumn": "dob"},
                        {
                            "col": "D",
                            "mode": "interpret",
                            "label": "Reason not titrated",
                            "sourceTable": "clinician_notes",
                            "sourceColumn": "note_text",
                            "sourceQuery": "SELECT mrn, note_text FROM clinician_notes ORDER BY mrn",
                            "instruction": "Extract reason if documented.",
                        },
                    ],
                },
                "query": {"sql": "SELECT mrn, dob FROM patients ORDER BY mrn", "parameters": []},
            })
            self.assertEqual(proc.returncode, 0, proc.stderr)
            payload = json.loads(proc.stdout)
            self.assertTrue(payload["ok"], payload)
            # Only direct cells: 3 rows × 2 direct fields = 6
            self.assertEqual(payload["cellsWritten"], 6)
            self.assertEqual(payload["firstDataRow"], 4)
            self.assertEqual(payload["rowIds"], ["M001", "M002", "M003"])

            pending = payload["pendingInterpretFields"]
            self.assertEqual(len(pending), 1)
            entry = pending[0]
            self.assertEqual(entry["col"], "D")
            self.assertEqual(entry["sourceTable"], "clinician_notes")
            self.assertEqual(entry["sourceColumn"], "note_text")
            self.assertEqual(entry["firstDataRow"], 4)
            self.assertEqual(entry["lastDataRow"], 10)
            self.assertEqual(entry["instruction"], "Extract reason if documented.")
            self.assertEqual(entry["label"], "Reason not titrated")

            wb = load_workbook(workbook_path, data_only=True)
            patients = wb["Patients"]
            # Direct cells written with literal values
            self.assertEqual(patients["A4"].value, "M001")
            self.assertEqual(patients["B4"].value, "1965-04-12")
            # Interpret column left empty (filled later via table_write_values)
            self.assertIsNone(patients["D4"].value)

    def test_all_interpret_region_returns_row_mapping_without_writing(self):
        with tempfile.TemporaryDirectory() as directory:
            db_path = Path(directory) / "emr.sqlite"
            workbook_path = Path(directory) / "audit.xlsx"
            make_db(db_path)
            make_workbook(workbook_path)

            proc = run_populate({
                "workbookPath": str(workbook_path),
                "databasePath": str(db_path),
                "runId": "test-run",
                "region": {
                    "regionId": "R1",
                    "sheet": "Patients",
                    "kind": "row_per_entity",
                    "anchor": {"dataRange": "A4:D10", "rowIdColumn": "mrn"},
                    "fields": [
                        {
                            "col": "D",
                            "mode": "interpret",
                            "sourceTable": "notes",
                            "sourceColumn": "text",
                        },
                    ],
                },
                "query": {"sql": "SELECT mrn FROM patients ORDER BY mrn", "parameters": []},
            })
            self.assertEqual(proc.returncode, 0, proc.stderr)
            payload = json.loads(proc.stdout)
            self.assertTrue(payload["ok"], payload)
            self.assertEqual(payload["cellsWritten"], 0)
            self.assertEqual(payload["rowIds"], ["M001", "M002", "M003"])
            self.assertEqual(len(payload["pendingInterpretFields"]), 1)

            workbook = load_workbook(workbook_path)
            self.assertNotIn("Evidence", workbook.sheetnames)

    def test_scalar_kpi_writes_single_row(self):
        with tempfile.TemporaryDirectory() as directory:
            db_path = Path(directory) / "emr.sqlite"
            workbook_path = Path(directory) / "audit.xlsx"
            make_db(db_path)
            make_workbook(workbook_path)

            proc = run_populate({
                "workbookPath": str(workbook_path),
                "databasePath": str(db_path),
                "runId": "test-run",
                "region": {
                    "regionId": "R2",
                    "sheet": "Summary",
                    "kind": "scalar_kpi",
                    "fields": [
                        {"cell": "B3", "resultColumn": "total", "sourceTable": "patients", "sourceColumn": "mrn"},
                    ],
                },
                "query": {"sql": "SELECT COUNT(*) AS total FROM patients", "parameters": []},
            })
            self.assertEqual(proc.returncode, 0, proc.stderr)
            payload = json.loads(proc.stdout)
            self.assertTrue(payload["ok"], payload)
            self.assertEqual(payload["cellsWritten"], 1)


if __name__ == "__main__":
    unittest.main()
