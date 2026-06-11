import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
TABLE_TOOL = ROOT / ".opencode" / "tools" / "table.py"


def run_describe(workbook_path: Path):
    proc = subprocess.run(
        [sys.executable, str(TABLE_TOOL), json.dumps({"action": "describe_layout", "inputPath": str(workbook_path)})],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=False,
    )
    return proc


class DescribeLayoutTest(unittest.TestCase):
    def test_sparse_cells_empty_blocks_and_merged_ranges(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "template.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Patients"
            sheet["A1"] = "Diabetes Audit 2025"
            sheet.merge_cells("A1:F1")
            sheet["A3"] = "MRN"
            sheet["B3"] = "DOB"
            sheet["C3"] = "HbA1c"
            sheet["C4"].number_format = "0.0"
            for row in range(5, 11):
                for col in range(1, 7):
                    sheet.cell(row=row, column=col).value = None
            workbook.save(workbook_path)

            proc = run_describe(workbook_path)
            self.assertEqual(proc.returncode, 0, proc.stderr)
            payload = json.loads(proc.stdout)
            self.assertEqual(len(payload["sheets"]), 1)
            sheet_layout = payload["sheets"][0]
            self.assertEqual(sheet_layout["name"], "Patients")
            self.assertIn("A1:F1", sheet_layout["mergedRanges"])

            cell_map = {entry["cell"]: entry for entry in sheet_layout["cells"]}
            self.assertEqual(cell_map["A1"]["value"], "Diabetes Audit 2025")
            self.assertEqual(cell_map["A3"]["value"], "MRN")
            self.assertEqual(cell_map["C4"].get("numberFormat"), "0.0")
            self.assertNotIn("A2", cell_map)

            empty_ranges = {block["range"] for block in sheet_layout["emptyBlocks"]}
            self.assertTrue(any(r.startswith("A2:") for r in empty_ranges))

    def test_rejects_non_xlsx(self):
        with tempfile.TemporaryDirectory() as directory:
            csv_path = Path(directory) / "x.csv"
            csv_path.write_text("a,b\n1,2\n")
            proc = run_describe(csv_path)
            self.assertNotEqual(proc.returncode, 0)
            self.assertIn("XLSX", proc.stderr)


if __name__ == "__main__":
    unittest.main()
