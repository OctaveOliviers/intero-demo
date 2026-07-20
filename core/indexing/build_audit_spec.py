"""Builder for the audit specification (`spec.json`) — T11 rewrite.

**The altitude split (mirrors `build_database_model`):** the `fields[]`
skeleton is extracted MECHANICALLY from the workbook — every non-empty
row-1 header cell on every sheet becomes one field with its `number`,
`section` (the sheet), `cell` (column letter), verbatim `name`, and the
FK-stable `id` (`{slugify(sheet)}/{slugify(header)}` when multi-sheet, the
bare slug otherwise — `build_populate_spec._field_slug`'s convention, same
`core.slug.slugify` both sides). One LLM call supplies ONLY the clinical
judgment: per-field `type`/`unit`/`format`/`permitted_values`/`notes`
(joined back by `number`), plus the audit-level `title`/`description`/
`grain`/`version`/`deadline`/section display names and the ~5 suggested
`inclusion_criteria`. The LLM can never add, drop, rename, or move a field
— structure is not its to decide.

This kills the A5 bug list wholesale (BUILD-PLAN §A5): #2 (`ref.endswith("1")`
matched A11/A21 as headers — rows are now matched structurally), #4
(`slugify(header) ≠ slugify(name)` — `name` IS the verbatim header), and
#3/#7/#8 (the coverage guard, its truncated retry feedback, and the whole
LLM-enumeration altitude — there is nothing to cover-check when the model
cannot drop a field).

The field spec is REGENERABLE and database-agnostic — built before any
database is chosen, so it carries NO `kind` (direct/interpret is decided at
mapping). User-set state — `inclusion_criteria[].default`, library-added
`notes`/`permitted_values`, a shortened `field.id`, the `deadline` — is
PRESERVED across re-index by a small merge on stable keys (`field.number` /
`criterion.id`, §6). The model is validated against the S0 schema
(`audit-spec.schema.json`) before it is returned, so the service never
writes a broken file.
"""

from __future__ import annotations

import json
import re
import warnings
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from core.clients import llm
from core.clients.llm_builder import (
    BuilderValidationError,
    build_validated,
    parse_json_object,
)
from core.contracts import validate_against_schema
from core.slug import slugify

_SCHEMA_FILE = "audit-spec.schema.json"
_VALID_TYPES = {"category", "number", "date", "text", "boolean"}

# "<column letters><row number>" — structural row matching (the old string
# suffix test was bug #2: ref.endswith("1") matched A11, A21, A101…).
_CELL_REF = re.compile(r"^([A-Z]+)(\d+)$")
_ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


_PROMPT = """\
You annotate the specification of a clinical audit. The audit's field STRUCTURE
has already been extracted mechanically from its workbook — you are given the
final numbered field list (number, sheet, cell, verbatim header) plus the raw
layout for context. You CANNOT add, drop, rename, renumber, or move fields.
Produce ONLY a single JSON object — no preamble, no markdown fences, no
commentary.

You supply the CLINICAL JUDGMENT: what each field means and expects, and the
audit-level prose. The spec is database-agnostic — you do NOT know which
database will source the values, so you MUST NOT name any table or column, and
you MUST NOT say whether a field is copied or interpreted (that is decided
later, at mapping).

Output shape:

{
  "title": "<human-readable audit title>",
  "description": "<one paragraph: what this audit measures and its provenance>",
  "version": "<dataset/spec version if known, else omit>",
  "grain": "<what one row represents, e.g. 'one row per baby / birth record'>",
  "deadline": "<the audit's submission deadline as an ISO date YYYY-MM-DD, ONLY if the workbook layout states one; else omit>",
  "sections": [ { "id": "<sheet name, EXACTLY as given>", "name": "<descriptive display name for that sheet's section>" } ],
  "fields": [
    { "number": <the field's given number — your join key>,
      "type": "category | number | date | text | boolean",
      "unit": "<unit if numeric, optional>",
      "format": "<entry format hint, optional>",
      "permitted_values": { "<code>": "<meaning>" },
      "notes": "<operational guidance + rationale + any standard citation inline, e.g. [NG18: 1.2.46]>" }
  ],
  "inclusion_criteria": [
    { "id": "<slug>", "label": "<human-readable label>",
      "type": "category | number | date | boolean" }
  ]
}

Rules:
- Emit one `fields` entry for EVERY number in the given field list, keyed by
  that exact `number`. Entries for unknown numbers are discarded.
- `type` is exactly one of: category, number, date, text, boolean.
  * `number` / `date` from the header's clinical meaning (e.g. "Apgar at 1 min"
    → number, "Date of birth" → date).
  * `boolean` for a presence/normality question answered yes/no — "…done?",
    "…given?", "Admitted to ICU", "Previous caesarean". Most short clinical
    yes/no headers are boolean even without a question mark.
  * `category` when the field takes one value from a fixed coded set you can
    state — the workbook shows it, or it is a well-established set from the
    audit's domain (e.g. mode-of-delivery codes). Emit the set in
    `permitted_values`.
  * `text` for free entry — a finding description, a comorbidity list, a drug
    list, a ward name. When genuinely unsure between category and text, use
    text.
- `permitted_values` is the audit's canonical CODED set as a code -> meaning map
  (e.g. {"1":"Male","2":"Female"}). It MUST accompany every `category` field —
  if you cannot state the fixed set, the field is not a category. NEVER invent
  codes for sets you do not actually know: a wrong coded set silently corrupts
  every later mapping. Never write it as prose.
- `notes` is the ONE prose field — fold guidance, rationale, and any standard
  citation into it. Write a useful note for every field (what is recorded, how,
  any edge rules); a mapper who has never seen the template must understand the
  field from the note alone.
- `sections` carries a display name for every sheet in the given structure —
  keep `id` EXACTLY the sheet name you were given.
- `inclusion_criteria`: suggest only the ~5 MOST LIKELY dimensions a clinician
  would filter this audit's cohort on. Anchor each one to a dimension the
  template's OWN fields record, prioritising the audit's PRIMARY clinical
  measurements — the exposure, the headline outcome scores, the key event/
  admission flags — over administrative attributes (e.g. for a paediatric
  diabetes audit: diabetes type, age at visit, HbA1c, hospital admissions —
  not GP practice codes or postcode). Each is a database-agnostic
  CONCEPT — give it an id, a label, and a type. Do NOT pick default values;
  defaults are set later.
- Do NOT name any database table or column anywhere.
- Output only the JSON object, starting with `{`.
"""


# --- workbook layout extraction (deterministic skeleton) ----------------------

def _serialize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _describe_sheet(worksheet: Any) -> dict[str, Any]:
    max_row = worksheet.max_row or 0
    max_col = worksheet.max_column or 0
    cells: list[dict[str, Any]] = []
    empty_row_flags: list[bool] = []

    for row_index in range(1, max_row + 1):
        row_fully_empty = True
        for col_index in range(1, max_col + 1):
            cell = worksheet.cell(row=row_index, column=col_index)
            value = _serialize_cell(cell.value)
            number_format = (
                cell.number_format
                if cell.number_format and cell.number_format != "General"
                else None
            )
            if value != "" and value is not None:
                row_fully_empty = False
                entry: dict[str, Any] = {"cell": cell.coordinate, "value": value}
                if number_format:
                    entry["numberFormat"] = number_format
                cells.append(entry)
            elif number_format:
                cells.append({"cell": cell.coordinate, "value": None, "numberFormat": number_format})
        empty_row_flags.append(row_fully_empty)

    empty_blocks: list[str] = []
    if max_col > 0:
        last_col_letter = get_column_letter(max_col)
        start: int | None = None
        for index, is_empty in enumerate(empty_row_flags, start=1):
            if is_empty and start is None:
                start = index
            elif not is_empty and start is not None:
                empty_blocks.append(f"A{start}:{last_col_letter}{index - 1}")
                start = None
        if start is not None:
            empty_blocks.append(f"A{start}:{last_col_letter}{max_row}")

    merged_ranges = (
        [str(ref) for ref in worksheet.merged_cells.ranges]
        if hasattr(worksheet, "merged_cells")
        else []
    )

    return {
        "name": worksheet.title,
        "dimensions": {"rows": max_row, "cols": max_col},
        "mergedRanges": merged_ranges,
        "cells": cells,
        "emptyBlocks": empty_blocks,
    }


def extract_layout(workbook_path: Path) -> list[dict[str, Any]]:
    # openpyxl warns about unsupported extensions (e.g. Data Validation) it will
    # drop on read. We only read layout, never write back, so the dropped
    # extensions are irrelevant — silence the noise.
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=UserWarning)
        workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    return [_describe_sheet(workbook[name]) for name in workbook.sheetnames]


def extract_field_skeleton(
    sheets: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str]]:
    """The mechanical `fields[]` skeleton: one field per non-empty row-1 header
    cell, in workbook order. Returns ``(fields, sheet_names)`` where
    ``sheet_names`` lists the sheets that contributed at least one field.

    Each field carries the structure the LLM may never touch: `number` (1-based
    across the whole workbook), `section` (the sheet name — multi-sheet audits
    only), `cell` (column letter), verbatim `name`, and the FK-stable `id`
    (sheet-prefixed when multi-sheet; `build_populate_spec._field_slug`'s
    convention). Duplicate ids raise — two cells must not share a slug.
    """
    headers: list[tuple[str, str, str]] = []  # (sheet, column_letter, header)
    for sheet in sheets:
        for cell in sheet.get("cells", []):
            m = _CELL_REF.match(str(cell.get("cell") or ""))
            value = cell.get("value")
            if m and m.group(2) == "1" and isinstance(value, str) and value.strip():
                headers.append((sheet.get("name", "?"), m.group(1), value.strip()))

    sheet_names = sorted({s for s, _, _ in headers},
                         key=lambda n: [s.get("name") for s in sheets].index(n))
    multi_sheet = len(sheet_names) > 1

    fields: list[dict[str, Any]] = []
    seen: set[str] = set()
    for number, (sheet_name, column, header) in enumerate(headers, start=1):
        base = slugify(header) or f"field_{number}"
        fid = f"{slugify(sheet_name)}/{base}" if multi_sheet else base
        if fid in seen:
            raise BuilderValidationError(
                f"duplicate field id {fid!r} (cell {sheet_name}!{column}1, header "
                f"{header!r}); two fields cannot share a slug — rename one at source"
            )
        seen.add(fid)
        field: dict[str, Any] = {
            "id": fid,
            "number": number,
            "name": header,
            "cell": column,
        }
        if multi_sheet:
            field["section"] = sheet_name
        fields.append(field)
    return fields, sheet_names


# --- assembly + state preservation --------------------------------------------

def _prose_by_number(raw: Any) -> dict[int, dict[str, Any]]:
    """Index the LLM's per-field prose records by their `number` join key."""
    out: dict[int, dict[str, Any]] = {}
    for record in raw if isinstance(raw, list) else []:
        if not isinstance(record, dict):
            continue
        try:
            number = int(record.get("number"))
        except (TypeError, ValueError):
            continue
        out[number] = record
    return out


def _merge_field_prose(field: dict[str, Any], prose: dict[str, Any]) -> dict[str, Any]:
    """One spec field = the mechanical skeleton + the LLM's judgment for it.
    Structure (`id`/`number`/`name`/`cell`/`section`) is never the LLM's to
    write; a missing/invalid `type` falls back to `text` so the spec stays
    schema-valid (the eval's type_match shows the gap rather than the build
    dying on it)."""
    ftype = str(prose.get("type") or "").strip().lower()
    field["type"] = ftype if ftype in _VALID_TYPES else "text"
    for key in ("unit", "format", "notes"):
        value = str(prose.get(key) or "").strip()
        if value:
            field[key] = value
    pv = prose.get("permitted_values")
    if isinstance(pv, dict) and pv:
        field["permitted_values"] = {str(k): str(v) for k, v in pv.items()}
    elif field["type"] == "category":
        # A category WITHOUT its coded set is a guess (the prompt forbids it);
        # downgrade to text rather than ship a half-specified coded field.
        field["type"] = "text"
    return field


def _clean_criterion(raw: dict[str, Any]) -> dict[str, Any] | None:
    cid = str(raw.get("id") or "").strip()
    label = str(raw.get("label") or "").strip()
    ctype = str(raw.get("type") or "").strip().lower()
    if not cid or not label or ctype not in _VALID_TYPES:
        return None
    # A1 suggests these at upload with no default yet (§6); the LLM never invents
    # a default value.
    return {"id": cid, "label": label, "type": ctype, "suggested": True, "default": None}


def _assemble(
    audit_id: str,
    skeleton: list[dict[str, Any]],
    sheet_names: list[str],
    prose: dict[str, Any],
) -> dict[str, Any]:
    by_number = _prose_by_number(prose.get("fields"))
    fields = [
        _merge_field_prose(dict(field), by_number.get(field["number"], {}))
        for field in skeleton
    ]
    criteria = [c for c in (_clean_criterion(r) for r in prose.get("inclusion_criteria", []) or []
                            if isinstance(r, dict))
                if c is not None]
    model: dict[str, Any] = {
        "schema_version": "1",
        "audit": audit_id,
        "title": str(prose.get("title") or audit_id).strip(),
        "description": str(
            prose.get("description") or f"Audit specification for {audit_id}."
        ).strip(),
        "grain": str(prose.get("grain") or "one row per record").strip(),
        "fields": fields,
        "inclusion_criteria": criteria,
    }
    version = str(prose.get("version") or "").strip()
    if version:
        model["version"] = version
    deadline = str(prose.get("deadline") or "").strip()
    if _ISO_DATE.match(deadline):
        model["deadline"] = deadline
    # Sections are the sheets (deterministic ids); the LLM supplies only the
    # display names. Single-sheet audits carry no sections block.
    if len(sheet_names) > 1:
        display = {
            str(s.get("id") or "").strip(): str(s.get("name") or "").strip()
            for s in prose.get("sections", []) or []
            if isinstance(s, dict)
        }
        model["sections"] = [
            {"id": name, "name": display.get(name) or name} for name in sheet_names
        ]
    return model


def merge_preserved_state(new_model: dict[str, Any], old_model: dict[str, Any] | None) -> dict[str, Any]:
    """Preserve user-set state across re-index (§6). The regenerated field spec is
    authoritative for structure, but durable user state is carried forward by
    stable key: `inclusion_criteria[].default` by `id`; library-added
    `notes`/`permitted_values` + the **stable `field.id`** by `field.number`;
    and a hand-set `deadline` when the regenerated spec found none. `id` is
    preserved so a shortened library-set slug (e.g. `delivery` for "Mode of
    delivery") survives a re-index — the FK that downstream `mapping.json`
    executables already hold-onto cannot be broken by a rename. `notes` /
    `permitted_values` are kept only where the regenerated field left them empty,
    so a fresh spec never blows away a clinical lead's fill."""
    if not old_model:
        return new_model

    old_defaults = {
        c.get("id"): c.get("default")
        for c in old_model.get("inclusion_criteria", []) or []
        if isinstance(c, dict) and c.get("default") is not None
    }
    for crit in new_model.get("inclusion_criteria", []):
        preserved = old_defaults.get(crit["id"])
        if preserved is not None:
            crit["default"] = preserved

    old_fields = {
        f.get("number"): f
        for f in old_model.get("fields", []) or []
        if isinstance(f, dict) and isinstance(f.get("number"), int)
    }
    for field in new_model.get("fields", []):
        old = old_fields.get(field["number"])
        if not old:
            continue
        old_id = str(old.get("id") or "").strip()
        if old_id:
            field["id"] = old_id
        if not field.get("notes") and old.get("notes"):
            field["notes"] = old["notes"]
        if not field.get("permitted_values") and old.get("permitted_values"):
            field["permitted_values"] = old["permitted_values"]

    if not new_model.get("deadline") and old_model.get("deadline"):
        new_model["deadline"] = old_model["deadline"]
    return new_model


def _render_skeleton(fields: list[dict[str, Any]]) -> str:
    """The given field structure, rendered compactly for the prompt."""
    lines = ["FIELD STRUCTURE (final — annotate by `number`, never restructure):"]
    for f in fields:
        section = f" sheet={f['section']!r}" if f.get("section") else ""
        lines.append(f"  {f['number']}.{section} cell={f['cell']!r} header={f['name']!r}")
    return "\n".join(lines)


async def build_audit_spec(
    workbook_path: Path,
    audit_id: str,
    *,
    display_name: str | None = None,
    previous: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Build a validated `spec.json` document (a dict). The `fields[]` skeleton is
    mechanical (`extract_field_skeleton`); one LLM call supplies the per-field
    judgment + audit-level prose + suggested inclusion criteria; deterministic
    code merges them (structure always wins), merges preserved user state
    (`previous`), and validates against the S0 schema. Retries on malformed LLM
    output; raises `BuilderValidationError` when the attempts run out."""
    sheets = extract_layout(workbook_path)
    skeleton, sheet_names = extract_field_skeleton(sheets)
    if not skeleton:
        raise BuilderValidationError(
            f"no header cells found in {workbook_path.name} — nothing to specify"
        )
    hint = display_name or workbook_path.stem
    user_input = (
        f"filename_hint: {hint}\n\n"
        + _render_skeleton(skeleton)
        + "\n\nRAW LAYOUT (context only):\n"
        + json.dumps({"sheets": sheets}, ensure_ascii=False, indent=2)
    )

    async def _attempt(instructions: str) -> tuple[dict[str, Any] | None, list[str]]:
        # The prose pass dominates output size — every field gets a type +
        # note (+ permitted_values for coded fields), ~80 tokens each. 12k
        # carries ~150 fields comfortably.
        raw = await llm.respond(
            instructions, user_input,
            max_output_tokens=12000, temperature=0.0, stage="index_audit",
        )
        try:
            prose = parse_json_object(raw)
        except ValueError as exc:
            return None, [f"LLM output was not parseable JSON: {exc}"]
        model = merge_preserved_state(
            _assemble(audit_id, skeleton, sheet_names, prose), previous
        )
        problems = validate_against_schema(model, _SCHEMA_FILE)
        if not problems:
            # The structure cannot be wrong (it is mechanical); hold the LLM
            # to its one job instead: every field annotated.
            unannotated = [
                f["number"] for f in model["fields"]
                if f["number"] not in _prose_by_number(prose.get("fields"))
            ]
            if unannotated:
                problems = [
                    "fields missing an annotation record (join by `number`): "
                    + ", ".join(str(n) for n in unannotated)
                ]
        return model, problems

    return await build_validated(_attempt, prompt=_PROMPT, label="audit spec")
