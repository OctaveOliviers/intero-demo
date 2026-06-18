"""Tests for building a workbook purely from state-DB cells.

The spine fills cells into the state DB and streams them to the live grid; there
is no on-disk workbook. build_cell_workbook reconstructs the grid from the cells'
SHEET!A1 refs — these tests pin that values land on their refs verbatim (no
numeric coercion, so identifiers survive), headers come from the field-name map,
pending (value-less) rows still materialise, and malformed refs are skipped
rather than raising.
"""

from __future__ import annotations

import io
import unittest
from types import SimpleNamespace

from openpyxl import load_workbook

from core.running.workbook_stream import build_cell_workbook, sheets_from_cells


def _cell(ref: str, value, field: str | None = None):
    return SimpleNamespace(ref=ref, value=value, field=field)


def _read(wb, sheet="ALL"):
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    reread = load_workbook(out, data_only=True)
    try:
        ws = reread[sheet]
        return list(ws.iter_rows(min_row=1, max_col=2, values_only=True))
    finally:
        reread.close()


class BuildCellWorkbookTest(unittest.TestCase):
    def test_values_land_on_their_refs_verbatim(self):
        wb = build_cell_workbook([
            _cell("ALL!A2", "Alice", field="member"),
            _cell("ALL!B2", "42", field="value"),
            _cell("ALL!A3", "Bob", field="member"),
            _cell("ALL!B3", "3.5", field="value"),
        ])
        rows = _read(wb)
        # Every value stays a string — no numeric coercion, so identifiers like
        # NHS numbers ("00456", 10-digit codes) survive verbatim.
        self.assertEqual(rows[1], ("Alice", "42"))
        self.assertEqual(rows[2], ("Bob", "3.5"))

    def test_identifier_with_leading_zeros_is_preserved(self):
        wb = build_cell_workbook([_cell("ALL!A2", "00456", field="member")])
        self.assertEqual(_read(wb)[1][0], "00456")

    def test_header_row_uses_field_display_names_with_slug_fallback(self):
        wb = build_cell_workbook(
            [
                _cell("ALL!A2", "Alice", field="nhs_number"),
                _cell("ALL!B2", "x", field="unmapped_field"),
            ],
            field_names={"nhs_number": "NHS Number"},
        )
        rows = _read(wb)
        # Mapped field -> display name; unmapped -> the raw slug.
        self.assertEqual(rows[0], ("NHS Number", "unmapped_field"))

    def test_pending_cells_materialise_their_rows(self):
        # Value-less (pending) cells still produce a row so the blank cohort grid
        # renders before any tier fills a value. The "" anchor reads back as None
        # after a save/reload, but the row itself is materialised — that's the
        # point: iter_rows yields a tuple for every cohort row.
        wb = build_cell_workbook([
            _cell("ALL!A2", None, field="member"),
            _cell("ALL!A3", None, field="member"),
        ])
        rows = _read(wb)
        self.assertEqual(len(rows), 3)  # header + 2 cohort rows
        self.assertIn(rows[1][0], (None, ""))
        self.assertIn(rows[2][0], (None, ""))

    def test_malformed_refs_are_skipped_not_raised(self):
        # A single bad ref must not raise (and 500 the whole export).
        wb = build_cell_workbook([
            _cell("ALL!A2", "ok", field="member"),
            _cell("GHOST", "x"),            # no sheet separator
            _cell("ALL!", "y"),             # empty A1
            _cell("ALL!A0", "z"),           # row 0 is invalid
            _cell("ALL!foo", "z"),          # no row number
        ])
        rows = _read(wb)
        self.assertEqual(rows[1][0], "ok")

    def test_sheets_from_cells_shape(self):
        sheets = sheets_from_cells([
            _cell("ALL!A2", "Alice", field="member"),
            _cell("DETAIL!A2", "Alice", field="member"),
        ])
        names = [s["name"] for s in sheets]
        self.assertEqual(names, ["ALL", "DETAIL"])  # deterministic (sorted)
        self.assertIn("columns", sheets[0]["meta"])


if __name__ == "__main__":
    unittest.main(verbosity=2)
