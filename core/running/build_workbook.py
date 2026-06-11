"""Build an initial run workbook from an audit-spec + a known cohort size.

The storage contract (storage-layout.md §3) says ``workbook.xlsx`` exists only
when an audit was uploaded as a workbook, and that the system never reads it
at run time. For audits created any other way (seeded fixtures, library
templates indexed from a database model, etc.) there is no stored workbook;
the runtime materialises the empty result workbook from ``spec.json``.

This module is that materialiser. It produces an .xlsx with:

* One sheet per ``spec.sections[*].id`` — sheet names match section ids so the
  region/cell refs the orchestrator produces (e.g. ``ALL!T2``) resolve.
* Row 1 = column headers, placed at ``field.cell`` for every field whose
  ``section`` belongs to that sheet.
* Rows 2..(cohort_size + 1) = empty rows pre-allocated so the front-end's
  ``cell_update`` writes (which target rows by 1-based index) have a row to
  land in. The frontend silently drops writes to rows that don't exist in the
  workbook's ``data`` array, so pre-allocation is what makes the live-fill
  visible.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook


def build_initial_workbook(spec: dict[str, Any], cohort_size: int, dest: Path) -> None:
    """Write a fresh xlsx at ``dest`` with headers from ``spec`` and
    ``cohort_size`` empty data rows (one per cohort member)."""
    sections = spec.get("sections") or []
    if not sections:
        raise ValueError("spec has no sections; cannot build a workbook")
    fields = spec.get("fields") or []
    fields_by_section: dict[str, list[dict[str, Any]]] = {}
    for field in fields:
        section_id = field.get("section")
        if not section_id:
            continue
        fields_by_section.setdefault(section_id, []).append(field)

    wb = Workbook()
    # openpyxl creates one default sheet; reuse it for the first section.
    default_sheet = wb.active
    first = True
    for section in sections:
        section_id = section.get("id")
        if not section_id:
            continue
        if first:
            ws = default_sheet
            ws.title = section_id
            first = False
        else:
            ws = wb.create_sheet(title=section_id)
        # Row 1: headers at field.cell.
        for field in fields_by_section.get(section_id, []):
            cell_ref = field.get("cell")
            if not cell_ref:
                continue
            ws[f"{cell_ref}1"] = field.get("name") or field.get("id") or ""
        # Rows 2..N+1: pre-allocate so the frontend's cell_update writes have
        # a row to land in. openpyxl tracks max_row by what's been written, so
        # writing None does not bump max_row — we materialise a single cell at
        # the bottom-left corner of each data row with an empty string. The
        # frontend renders "" as a blank, so this is invisible to users; what
        # matters is that iter_rows() yields N+1 tuples, giving cell_update a
        # legitimate target index for every cohort member.
        for r in range(2, 2 + cohort_size):
            ws.cell(row=r, column=1, value="")

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    wb.close()
