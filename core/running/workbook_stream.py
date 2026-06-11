"""Workbook serialization helpers for run-stream events."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def read_workbook_sheets(result_path: Path) -> list[dict]:
    """Read `result.xlsx` into the `workbook_created.sheets` shape."""
    wb = load_workbook(result_path, data_only=True)
    try:
        sheets: list[dict] = []
        for ws in wb.worksheets:
            data = [
                [cell if cell is not None else None for cell in row]
                for row in ws.iter_rows(min_row=1, values_only=True)
            ]
            max_col = ws.max_column or (len(data[0]) if data else 0)
            columns = []
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                width = ws.column_dimensions[col_letter].width
                col_meta = {}
                if isinstance(width, (int, float)) and width > 0:
                    col_meta["width"] = int(round(width))
                columns.append(col_meta)
            sheets.append({
                "name": ws.title,
                "data": data,
                "meta": {"columns": columns},
            })
        return sheets
    finally:
        wb.close()
