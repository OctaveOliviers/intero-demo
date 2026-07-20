"""Build a table-population workbook purely from its state-DB cells.

state.db is the single source of truth: the two population steps write every
cell value there and stream it to the live grid, but nothing is ever written
back to disk. The cells are self-describing — each carries its ``SHEET!A1`` ref
(sheet + position) and its value — so the sheets, columns and rows can be
reconstructed from the refs alone. There is no on-disk ``result.xlsx`` and no
separate layout to drift against: a rebuilt workbook can only ever match what
table population actually wrote.

The one thing a cell does not carry is a human-readable column header (its
``field`` is a slug like ``nhs_number``); ``field_names`` maps that to the spec's
display name when available, falling back to the slug. Headers are cosmetic — a
missing/changed name can never misplace a value.
"""

from __future__ import annotations

import io
import re
from typing import Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# A well-formed A1 ref: column letters + a 1-based row. A stray ref ("ALL!",
# "ALL!A0", "ALL!foo") is skipped rather than raising.
_A1_RE = re.compile(r"^([A-Za-z]+)([1-9][0-9]*)$")

# Headers occupy row 1; data rows are 2+. Pre-allocated rows materialise their
# left edge with this so iter_rows() yields a tuple for every cohort row (the FE
# renders "" as blank). Matches the live grid.
_ROW_ANCHOR = ""


def _split_ref(ref: str):
    """``"ALL!B2"`` -> ``("ALL", "B2", "B", 2)``; ``None`` for a malformed ref."""
    sheet, sep, a1 = ref.partition("!")
    if not sep:
        return None
    match = _A1_RE.match(a1)
    if match is None:
        return None
    return sheet, a1, match.group(1).upper(), int(match.group(2))


def build_cell_workbook(cells: Iterable, field_names: dict | None = None) -> Workbook:
    """An openpyxl workbook reconstructed from ``cells``, in place of any on-disk
    layout. Sheets/columns/rows come from the refs; values are written verbatim
    (no numeric coercion, so clinical identifiers keep their leading zeros and
    full precision); row 1 is the per-column header from ``field_names``.

    Pending cells (value ``None``) still materialise their row, so the blank
    cohort grid renders before any value lands.
    """
    field_names = field_names or {}
    parsed = []
    sheet_order: list[str] = []
    for cell in cells:
        split = _split_ref(cell.ref)
        if split is None:
            continue
        sheet, a1, col, row = split
        if sheet not in sheet_order:
            sheet_order.append(sheet)
        parsed.append((cell, sheet, a1, col, row))
    # Sort so the run-start emit (built from precomputed pending cells, in region
    # order) and a later reload (built from store cells, ordered by ref) agree.
    sheet_order.sort()

    wb = Workbook()
    default = wb.active
    if not sheet_order:
        # No cells: a single empty sheet, nothing to populate.
        return wb

    ws_by_sheet: dict = {}
    for index, sheet in enumerate(sheet_order):
        ws = default if index == 0 else wb.create_sheet()
        ws.title = sheet
        ws_by_sheet[sheet] = ws

    header_field: dict[str, dict[str, str]] = {sheet: {} for sheet in sheet_order}
    rows_seen: dict[str, set[int]] = {sheet: set() for sheet in sheet_order}
    for cell, sheet, a1, col, row in parsed:
        ws = ws_by_sheet[sheet]
        if cell.value is not None:
            ws[a1] = cell.value
        header_field[sheet].setdefault(col, cell.field)
        rows_seen[sheet].add(row)

    for sheet in sheet_order:
        ws = ws_by_sheet[sheet]
        for col, field in header_field[sheet].items():
            ws[f"{col}1"] = field_names.get(field, field) if field else None
        for row in rows_seen[sheet]:
            if ws.cell(row=row, column=1).value is None:
                ws.cell(row=row, column=1, value=_ROW_ANCHOR)
    return wb


def workbook_to_sheets(wb: Workbook) -> list[dict]:
    """An openpyxl workbook -> the ``workbook_created.sheets`` shape (grid values
    + per-column widths) the SSE stream and ``/workbook`` endpoint emit."""
    sheets: list[dict] = []
    for ws in wb.worksheets:
        data = [
            [cell if cell is not None else None for cell in row]
            for row in ws.iter_rows(min_row=1, values_only=True)
        ]
        max_col = ws.max_column or (len(data[0]) if data else 0)
        columns = []
        for col_idx in range(1, max_col + 1):
            width = ws.column_dimensions[get_column_letter(col_idx)].width
            col_meta: dict = {}
            if isinstance(width, (int, float)) and width > 0:
                col_meta["width"] = int(round(width))
            columns.append(col_meta)
        sheets.append({"name": ws.title, "data": data, "meta": {"columns": columns}})
    return sheets


def sheets_from_cells(cells: Iterable, field_names: dict | None = None) -> list[dict]:
    """Convenience: ``cells`` -> the ``workbook_created.sheets`` shape."""
    return workbook_to_sheets(build_cell_workbook(cells, field_names))


def cells_to_xlsx(
    cells: Iterable,
    field_names: dict | None = None,
    inactive_rows_by_sheet: dict[str, list[int]] | None = None,
) -> io.BytesIO:
    """``cells`` -> a downloadable .xlsx byte stream.

    ``inactive_rows_by_sheet`` maps a sheet to the 1-based rows to drop (deselected
    cohort members), each list sorted descending so deletes don't shift later
    indices.
    """
    wb = build_cell_workbook(cells, field_names)
    if inactive_rows_by_sheet:
        for ws in wb.worksheets:
            for row in inactive_rows_by_sheet.get(ws.title, []):
                ws.delete_rows(row, 1)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
