import asyncio
import io
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
from starlette.responses import StreamingResponse

from core.store import Cell, Run, RunMember, Store
from server.routes import workbook as workbook_route


async def _read_streaming_body(response: StreamingResponse) -> bytes:
    data = b""
    async for chunk in response.body_iterator:
        data += chunk
    return data


class WorkbookDownloadPolicyTest(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.base = Path(self._tmp.name)
        self.db_path = self.base / "state.db"
        # No spec on disk -> headers fall back to the field slug. There is no
        # result.xlsx anywhere: the grid is built purely from state.db cells.
        self.audits_dir = self.base / "audits"

        store = Store(self.db_path)
        store.create_run(Run(id="run-1", audit_id="npda", status="completed"))
        store.upsert_run_member(
            RunMember(run_id="run-1", member="A", row_index=0, active=True)
        )
        store.upsert_run_member(
            RunMember(run_id="run-1", member="B", row_index=1, active=False)
        )
        store.upsert_run_member(
            RunMember(run_id="run-1", member="C", row_index=2, active=True)
        )
        # member -> (ALL!value, DETAIL!detail), keyed by data row.
        seed = {
            ("A", 2): (10, "alpha"),
            ("B", 3): (20, "beta"),
            ("C", 4): (30, "gamma"),
        }
        src = [{"database": "db", "query": "SELECT 1", "table_column": "t.c"}]
        for (member, row), (val, detail) in seed.items():
            store.upsert_cell(Cell(run_id="run-1", ref="ALL!A%d" % row, field="member",
                                   member=member, kind="direct", state="filled", value=member, sources=src))
            store.upsert_cell(Cell(run_id="run-1", ref="ALL!B%d" % row, field="value",
                                   member=member, kind="direct", state="filled", value=str(val), sources=src))
            store.upsert_cell(Cell(run_id="run-1", ref="DETAIL!A%d" % row, field="member",
                                   member=member, kind="direct", state="filled", value=member, sources=src))
            store.upsert_cell(Cell(run_id="run-1", ref="DETAIL!B%d" % row, field="detail",
                                   member=member, kind="direct", state="filled", value=detail, sources=src))
        store.close()

    def tearDown(self):
        self._tmp.cleanup()

    def test_download_excludes_inactive_members_but_keeps_history(self):
        with (
            patch.object(workbook_route, "AUDITS_DIR", self.audits_dir),
            patch.object(workbook_route, "Store", lambda: Store(self.db_path)),
        ):
            response = asyncio.run(workbook_route.download_workbook("run-1"))

        self.assertIsInstance(response, StreamingResponse)
        payload = asyncio.run(_read_streaming_body(response))
        wb = load_workbook(io.BytesIO(payload), data_only=True)
        ws = wb["ALL"]
        ws2 = wb["DETAIL"]
        rows = list(ws.iter_rows(min_row=1, max_col=2, values_only=True))
        rows2 = list(ws2.iter_rows(min_row=1, max_col=2, values_only=True))
        wb.close()

        # Values come from the state DB (built entirely from cells) and are kept
        # verbatim as strings — no numeric coercion. The header row is present and
        # inactive member B's row is dropped.
        self.assertEqual(rows, [
            ("member", "value"),
            ("A", "10"),
            ("C", "30"),
        ])
        self.assertEqual(rows2, [
            ("member", "detail"),
            ("A", "alpha"),
            ("C", "gamma"),
        ])

        store = Store(self.db_path)
        try:
            members = {m.member: m for m in store.list_run_members("run-1")}
        finally:
            store.close()
        self.assertIn("B", members)
        self.assertFalse(members["B"].active)

    def test_get_workbook_builds_full_grid_from_cells(self):
        with (
            patch.object(workbook_route, "AUDITS_DIR", self.audits_dir),
            patch.object(workbook_route, "Store", lambda: Store(self.db_path)),
        ):
            resp = asyncio.run(workbook_route.get_workbook("run-1"))

        by_name = {s.name: s for s in resp.sheets}
        # Reload builds the same populated grid as the download — values from the
        # state DB. get_workbook keeps ALL rows (history view); inactive filtering
        # is a download-only concern.
        self.assertEqual(by_name["ALL"].data[0], ["member", "value"])
        self.assertEqual(by_name["ALL"].data[1][:2], ["A", "10"])
        self.assertEqual(by_name["ALL"].data[2][:2], ["B", "20"])
        self.assertEqual(by_name["ALL"].data[3][:2], ["C", "30"])
        self.assertEqual(resp.runStatus, "completed")

    def test_download_404_when_run_missing(self):
        from fastapi import HTTPException
        with (
            patch.object(workbook_route, "AUDITS_DIR", self.audits_dir),
            patch.object(workbook_route, "Store", lambda: Store(self.db_path)),
        ):
            with self.assertRaises(HTTPException) as ctx:
                asyncio.run(workbook_route.download_workbook("nope"))
        self.assertEqual(ctx.exception.status_code, 404)


if __name__ == "__main__":
    unittest.main(verbosity=2)
