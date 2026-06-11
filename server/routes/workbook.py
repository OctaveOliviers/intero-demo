import io
import json
import re
from pathlib import Path

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from core.config import RUNS_DIR
from core.store import Store
from server.models import WorkbookResponse, SheetData, SheetMeta, SheetColumnMeta, CellMetadata

router = APIRouter()
_A1_ROW_RE = re.compile(r"^[A-Za-z]+([1-9][0-9]*)$")


def _row_from_a1(ref: str) -> int | None:
    match = _A1_ROW_RE.match(ref.strip())
    if match is None:
        return None
    return int(match.group(1))


def _inactive_member_rows_by_sheet(run_id: str) -> dict[str, list[int]]:
    store = Store()
    try:
        inactive_members = {
            member.member
            for member in store.list_run_members(run_id)
            if not member.active
        }
        if not inactive_members:
            return {}
        rows_by_sheet: dict[str, set[int]] = {}
        for cell in store.get_cells(run_id):
            if cell.member not in inactive_members:
                continue
            sheet, sep, a1 = cell.ref.partition("!")
            if not sep:
                continue
            row = _row_from_a1(a1)
            if row is None:
                continue
            rows_by_sheet.setdefault(sheet, set()).add(row)
        return {
            sheet: sorted(rows, reverse=True)
            for sheet, rows in rows_by_sheet.items()
            if rows
        }
    finally:
        store.close()


def _build_filtered_workbook_stream(
    workbook_path: Path,
    inactive_rows_by_sheet: dict[str, list[int]],
) -> io.BytesIO:
    wb = load_workbook(workbook_path)
    try:
        for ws in wb.worksheets:
            for row in inactive_rows_by_sheet.get(ws.title, []):
                ws.delete_rows(row, 1)
        out = io.BytesIO()
        wb.save(out)
    finally:
        wb.close()
    out.seek(0)
    return out


@router.get("/api/runs/{run_id}/workbook", response_model=WorkbookResponse)
async def get_workbook(run_id: str):
    run_dir = RUNS_DIR / run_id
    result_path = run_dir / "result.xlsx"
    metadata_path = run_dir / "metadata.json"

    if not result_path.exists():
        raise HTTPException(status_code=404, detail="Workbook not found.")

    try:
        wb = load_workbook(result_path, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read workbook: {e}")

    cell_metadata: dict[str, CellMetadata] = {}
    if metadata_path.exists():
        try:
            raw = json.loads(metadata_path.read_text(encoding="utf-8"))
            if isinstance(raw, dict):
                for cell_ref, entry in raw.items():
                    if isinstance(entry, dict) and entry.get("sql"):
                        evidence = entry.get("evidence")
                        cell_metadata[cell_ref] = CellMetadata(
                            value=entry.get("value"),
                            sql=entry["sql"],
                            explanation=entry.get("explanation"),
                            database=entry.get("database"),
                            evidence=evidence if isinstance(evidence, list) else None,
                        )
        except Exception:
            pass

    sheets: list[SheetData] = []
    for ws in wb.worksheets:
        data: list[list] = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            data.append([cell if cell is not None else None for cell in row])

        col_meta: list[SheetColumnMeta] = []
        for col_idx in range(1, ws.max_column + 1):
            dim = ws.column_dimensions.get(get_column_letter(col_idx))
            width = dim.width if dim and dim.width else None
            if isinstance(width, float) and width > 0:
                col_meta.append(SheetColumnMeta(width=int(round(width))))
            else:
                col_meta.append(SheetColumnMeta())
        sheets.append(SheetData(name=ws.title, data=data, meta=SheetMeta(columns=col_meta)))

    wb.close()

    return WorkbookResponse(sheets=sheets, cellMetadata=cell_metadata)


@router.get("/api/runs/{run_id}/download")
async def download_workbook(run_id: str):
    run_dir = RUNS_DIR / run_id
    result_path = run_dir / "result.xlsx"

    if not result_path.exists():
        raise HTTPException(status_code=404, detail="Workbook not found.")

    inactive_rows_by_sheet = _inactive_member_rows_by_sheet(run_id)
    if not inactive_rows_by_sheet:
        return FileResponse(
            path=str(result_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="result.xlsx",
            headers={"Content-Disposition": 'attachment; filename="result.xlsx"'},
        )

    try:
        stream = _build_filtered_workbook_stream(result_path, inactive_rows_by_sheet)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to build filtered workbook: {e}")

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="result.xlsx"'},
    )
